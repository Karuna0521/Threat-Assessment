from collections import OrderedDict
import random
import string
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer
from django.db import DatabaseError
from django.shortcuts import get_object_or_404, render
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.authtoken.models import Token
from django.contrib.auth import authenticate,login
from checklistApp.models import User
from .scripts.authentication import MyJWTAuthentication
from checklistApp.scripts.permissions import IsAdmin, IsReviewer
from checklistApp.serializers import *
from rest_framework import serializers
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework import generics
from datetime import date
import pandas as pd
import json
from rest_framework.parsers import MultiPartParser
from openpyxl import load_workbook


from django.contrib.sessions.models import Session
# from django.contrib.auth.models import Group  # Import the Group class

# Create your views here.
class UserRegistrationView(APIView):      
    def post(self, request):
        request_data = request.data
        # By default role is set to Auditor
        request_data['role']= "Auditor"    
        data = OrderedDict()
        data.update(request_data)
        try:
            if not request.data.get('key'):
                raise serializers.ValidationError("Key is required !")
            if not request.data.get('captcha_string'):
                raise serializers.ValidationError("captcha_string is required !")
            key = request.data["key"]
            ip = ''.join([char for char in key if not char.isalpha()])
            try:
                captcha = Captcha.objects.filter(key=key)
                if len(captcha) == 0:
                    raise serializers.ValidationError("key is invalid")
            except Captcha.DoesNotExist:
                raise serializers.ValidationError("key is invalid")
            if ip != request.META["REMOTE_ADDR"] and captcha[0].key != key:
                # self.captcha_update(captcha[0])
                captcha.update(count=captcha.count+1)
                raise serializers.ValidationError("IP mismatch")
            if captcha[0].captcha_string != str(data["captcha_string"]):
                if captcha[0].count == 3 :
                    print(captcha[0].captcha_string)
                    Captchas = Captcha.objects.filter(captcha_string=captcha[0].captcha_string)
                    Captchas.delete()
                    raise serializers.ValidationError("captcha was expired")
                captcha.update(count=captcha[0].count+1)
                raise serializers.ValidationError("Incorrect Captcha")
        except Captcha.DoesNotExist:
            raise serializers.ValidationError("captcha is invalid")
        try :
            user = User.objects.get(email=request.data["email"])
            return Response({'status':403, 'message':'Email address already exists..!! Please choose a different email.'}, status=status.HTTP_400_BAD_REQUEST)
        except User.DoesNotExist :
            ser = UserSerializer(data = request.data)
            if ser.is_valid():
                user = ser.save()
                # captcha.delete()
                # token, created = Token.objects.get_or_create(user=user)
                # return Response({'token': token.key,"status-code": "201","message":"User Registered successfully"}, status=status.HTTP_200_OK)       
                return Response({"status-code": "200","errors": [],"message":"User Registered successfully"}, status=status.HTTP_200_OK)       
        return Response({'status-code':403, 'message':'Something went wrong','errors':ser.errors}, status=status.HTTP_400_BAD_REQUEST)

class CaptchaStringAPIView(APIView):
    def generateCaptchaString(self):
        length=6
        charset=string.ascii_letters
        return ''.join(random.choice(charset) for _ in range(length))

    def post(self, request):
        captcha_str = self.generateCaptchaString()
        key = self.generateCaptchaString() + str(request.META["REMOTE_ADDR"]) + self.generateCaptchaString()
        captcha = CaptchaSerializer(data={"captcha_string": captcha_str, "key":key, "count":0})
        captcha.is_valid(raise_exception=True)
        c = captcha.save()
        print(str(c.id))
        return Response({'status-code': 200,
                         "errors": [],
                         'message':'Captcha Generated',
                         "data":{"c_id": str(c.id), "captcha_string": captcha_str, "key":key}}, status=status.HTTP_200_OK)

class MyTokenObtainPairSerializer(TokenObtainPairSerializer):
    @classmethod
    def get_token(cls, user):
        token = super().get_token(user)
        print(token)
        try : 
            Usergroup = User.objects.get(id=user.pk)
            print(Usergroup.pk)
        except User.DoesNotExist:  
            pass
        token["role"] = Usergroup.role
        token["user_id"] = user.pk
        token["full_name"] = user.full_name
        return token

class UserLogin(APIView):
    # authentication_classes=[MyJWTAuthentication]
    # permission_classes=[IsAuthenticated, IsAdmin]
    def post(self, request):
        data = OrderedDict()
        data.update(request.data)
        try:
            if not request.data.get('key'):
                raise serializers.ValidationError("Key is required !")
            if not request.data.get('captcha_string'):
                raise serializers.ValidationError("You need to complete the captcha!")
            key = request.data["key"]
            ip = ''.join([char for char in key if not char.isalpha()])
            try:
                captcha = Captcha.objects.filter(key=key)
                if len(captcha) == 0:
                    raise serializers.ValidationError("key is invalid")
            except Captcha.DoesNotExist:
                raise serializers.ValidationError("key is invalid")
            if ip != request.META["REMOTE_ADDR"] and captcha[0].key != key:
                # self.captcha_update(captcha[0])
                captcha.update(count=captcha.count+1)
                raise serializers.ValidationError("IP mismatch")
            if captcha[0].captcha_string != str(data["captcha_string"]):
                if captcha[0].count == 3 :
                    print(captcha[0].captcha_string)
                    Captchas = Captcha.objects.filter(captcha_string=captcha[0].captcha_string)
                    Captchas.delete()
                    raise serializers.ValidationError("captcha was expired")
                captcha.update(count=captcha[0].count+1)
                raise serializers.ValidationError("Incorrect Captcha. Please try again")
        except Captcha.DoesNotExist:
            raise serializers.ValidationError("captcha is invalid")

        email = request.data.get('email')
        password = request.data.get('password')
        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            user = None
        
        if user is not None:
            auth_user = authenticate(username=user.username, password=password)
            if auth_user is not None:
                
                login(request, auth_user)
                u = User.objects.get(id=auth_user.pk)
                refresh = MyTokenObtainPairSerializer.get_token(u)
                access_token = str(refresh.access_token)
                refresh_token = str(refresh)
                 # Get the session_key after login
                session_key = request.session.session_key
                print('session_key after login-->',session_key)
                
                return Response({
                    'msg': "User loggedIn Succesfully",
                    'refresh': refresh_token,
                    'access': access_token,
                    'session_key': session_key,
                })
        return Response({'error': 'Invalid credentials'}, status=status.HTTP_401_UNAUTHORIZED)

class UserLogout(APIView):
    def post(self, request):
        try:
            # Check if the request contains a 'session_id' in the JSON body
            session_id = request.data.get('session_id', None)

            # if not session_id:
            #     return Response({"detail": "Missing 'session_id' in the request body."}, status=status.HTTP_400_BAD_REQUEST)

            # # Get the current user's session key
            # current_session_key = request.session.session_key
            # print('current-->',current_session_key)
            # Check if the provided session key matches the current user's session key
            #if session_id != current_session_key:
            #    return Response({"detail": "Provided session key does not match the current user's session."}, status=status.HTTP_400_BAD_REQUEST)

            # Delete the session ID from the django_session table
            session = Session.objects.filter(session_key=session_id)
            if len(session) > 0:
                session.delete()
                return Response({"detail": "Successfully logged out."}, status=status.HTTP_200_OK)
            # else:
            #     return Response({"detail": "session not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({"detail": "Invalid request."}, status=status.HTTP_400_BAD_REQUEST)

class UserApi(APIView):
    authentication_classes=[MyJWTAuthentication]
    permission_classes = [IsAuthenticated, IsAdmin]
    
    def get(self, request,id=None):
       if id:
            user = User.objects.get(id=id)
            ser = UserSerializer(user)
            return Response({"payload": ser.data}, status=status.HTTP_200_OK)
       userData = User.objects.all()
       ser = UserSerializer(userData, many=True)
       return Response({'payload':ser.data})
    
    def patch(self, request):
        try:
            data = OrderedDict()
            data.update(request.data)
            user = User.objects.get(id=request.data['id'])
            print('User',user)
            
            new_role = request.data.get('role')
            print('new_role',new_role)
            old_role = user.role
            print('old_role',old_role)

            if new_role != old_role:
                try:
                    if new_role == 'Reviewer':
                        print('RRR Befoire')
                        audrev_mapping = AudRevMapping.objects.get(aud_id=user.id)
                        print('RRR After')
                    elif new_role == 'Auditor':
                        print('AAA Befoire')
                        audrev_mapping = AudRevMapping.objects.get(rev_id=user.id)
                        print('AAA After')
                    print('audrev_mapping',audrev_mapping)
                    audrev_mapping.delete()

                except AudRevMapping.DoesNotExist:
                    print('Except')
                    pass

            ser = UserSerializer(user, data=request.data, partial=True)
            if ser.is_valid():
                ser.save()
                return Response({'status-code': 200, "errors": [], 'message': 'User Data is updated'}, status=status.HTTP_200_OK)

            return Response({'status-code': 403, 'message': 'Something went wrong', 'errors': ser.errors}, status=status.HTTP_403_FORBIDDEN)
        except Exception as e:
            return Response({'status-code': 403, 'message': 'Invalid id in the url'}, status=status.HTTP_403_FORBIDDEN)

        
    def delete(self, request, id=None):
        try:
            # id = request.GET.get('id')
            user = User.objects.get(id =request.data['id'])
            # user_id.delete()
            data = {'status': 'inactive'} # setting status to inactive
            ser = UserSerializer(user, data = data, partial=True)
            
            if ser.is_valid():
                ser.save()
                try:
                    if user.role == 'Reviewer':
                        audrev_mapping = AudRevMapping.objects.get(rev_id=request.data['id'])
                    elif user.role == 'Auditor':
                        audrev_mapping = AudRevMapping.objects.get(aud_id=request.data['id'])

                    audrev_mapping.delete()

                except AudRevMapping.DoesNotExist:
                    pass
                # delete aud rev mapping for that user
                # audrev_mapping = dict()
                
                # if user.role == 'Reviewer': # if role rev then get mapping by rev_id
                #     audrev_mapping = AudRevMapping.objects.get(rev_id=request.data['id'])
                    
                # if user.role == 'Auditor': # if role aud then get mapping by aud_id
                #     audrev_mapping = AudRevMapping.objects.get(aud_id=request.data['id'])
                # audrev_mapping.delete()
                return Response({'status-code': 200,"errors": [],'message':'User is soft deleted(status=inactive)'},status=status.HTTP_200_OK)
            return Response({'status-code':403, 'message':'Something went wrong','errors':ser.errors})
        except Exception as e:
            return Response({'status-code':403,'message':'Inavlid id'},status=status.HTTP_403_FORBIDDEN)  

class AuditorData(APIView):
    # authentication_classes=[MyJWTAuthentication]
    # permission_classes = [IsAuthenticated, IsReviewer]
    def get(self, request):
        # auditor_data = get_object_or_404(User, id=auditor_id, role='auditor')
        auditor_data = User.objects.filter(role='Auditor')
        serializer = UserSerializer(auditor_data, many=True)
        print(auditor_data)
        return Response({'status-code': 200,"errors": [], 'payload':serializer.data},status=status.HTTP_200_OK)
    
    def post(self, request):   
        rev_id = request.data.get('rev_id')  # assuming rev_id is passed in the request body
        if not rev_id:
            return Response({'status-code': 400, 'errors': ['rev_id not provided'], 'payload': []}, status=status.HTTP_400_BAD_REQUEST)

        auditor_ids = AudRevMapping.objects.filter(rev_id=rev_id).values_list('aud_id', flat=True)
        auditors = User.objects.filter(id__in=auditor_ids, status='active')
        serializer = UserSerializer(auditors, many=True)
        auditors_data = serializer.data
        # auditors_data = self.get_auditors_mapped_to_reviewer(rev_id)

        return Response({'status-code': 200, "errors": [], 'payload': auditors_data}, status=status.HTTP_200_OK)
        
class OptionsView(APIView):
    def get(self, request):
        options = Options.objects.all()
        serializer = OptionsSerializer(options, many=True)
        return Response({'status-code':200, 'errors':[], 'payload':serializer.data},status=status.HTTP_200_OK)

    def post(self, request):
        option_text = str(request.data.get('option_text')).capitalize()
        existing = Options.objects.filter(option_text=option_text)
        if len(existing) > 0:
            return Response({'status-code':400, 'message': 'Option ' + option_text + ' already exists ..!'},status=status.HTTP_400_BAD_REQUEST)
            
        optionSer = OptionsSerializer(data={"option_text": option_text})
        optionSer.is_valid(raise_exception=True)
        o = optionSer.save()
        return Response({'status-code': 200,
                         "errors": [],
                         'message':'option created successfully ..!',
                         "data":{"id": str(o.id), "option_text": option_text}}, status=status.HTTP_200_OK)
    
    def delete(self, request):
        option_id = request.data.get('id')
        try:
            data = Options.objects.all()
            print(data)
            option_data = Options.objects.get(id=option_id)
            option_data.delete()
            return Response({'status-code':200,'errors':[], 'message':'Option '+str(option_data.option_text)+' deleted successfully'})
        except:
            raise serializers.ValidationError(str(option_id)+" is an Invalid id")


    def patch(self,request):
        option_data = request.data
        data = dict()
        data['id'] = option_data['id']
        data['option_text'] = str(option_data['option_text']).capitalize()

        option = Options.objects.filter(id=data['id'])
        if len(option) > 0:
            serializer = OptionsSerializer(option[0], data=data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response({'status-code':200, 'errors':[], 'message':'Option '+ str(option[0].option_text) +' updated successfully'},status=status.HTTP_200_OK)
            return Response({'status-code':400, 'message':'Invalid Option', 'errors': serializer.errors},status=status.HTTP_400_BAD_REQUEST)
        raise serializers.ValidationError(str(option_data['id'])+" is an Invalid id")

class AppInfoView(APIView):
    def post(self,request):
        app_info_data = request.data
        if not request.data.get('app_name'):
            raise serializers.ValidationError("Application name is required !")
        existing_app = AppInfo.objects.filter(app_name=app_info_data['app_name'])
        rev_id = app_info_data.get('reviewer_id')
        reviewer = User.objects.filter(id=rev_id)
        if len(reviewer)>0:
            if reviewer[0].role not in ['Reviewer' , 'reviewer']:
                raise serializers.ValidationError(str(rev_id) + " is not a Reviewer!")
        else:
            raise serializers.ValidationError("Reviewer Does Not exists with id:" + str(rev_id))
        if len(existing_app) > 0:
            return Response({'status-code':400, 'message': app_info_data['app_name'] + ' already exists ..!'},status=status.HTTP_400_BAD_REQUEST)
        if rev_id is not None:
            app_info_data['reviewer_assigned_date'] = date.today()
        serializer = AppInfoSerializer(data=app_info_data)
        if serializer.is_valid():
            serializer.save()
            return Response({'status-code':200, 'error':[],'message':'Application added successfully','payload':serializer.data})
        return Response({'status-code':400, 'message':'Invalid Option', 'errors': serializer.errors})
    
    def get(self, request):
        app_info = AppInfo.objects.all()
        serializer = AppInfoSerializer(app_info, many=True)
        return Response({'status-code':200, 'payload':serializer.data},status=status.HTTP_200_OK)

    def patch(self, request):
        app_info_data = request.data
        rev_id = app_info_data.get('reviewer_id')
        existing_app = AppInfo.objects.filter(app_name=app_info_data['app_name'])   
        if len(existing_app) > 0:
            existing_rev = existing_app[0].reviewer_id
            if rev_id != existing_rev:
                app_info_data['reviewer_assigned_date'] = date.today()
            serializer = AppInfoSerializer(existing_app[0], data = app_info_data, partial=True) 
            if serializer.is_valid():
                serializer.save()
                return Response({'status-code':200, 'errors':[], 'message':'Data updated Successfully..!'})
        else:
            raise serializers.ValidationError('Applicationdssda Does not exists..!')

    def delete(self,request):
        app_id = request.data.get('id')
        try:
            app_data = AppInfo.objects.get(id=app_id)
            app_data.delete()
            return Response({'status-code':200, 'errors':[],'message':'Application deleted successfully'})
        except:
            raise serializers.ValidationError(str(app_id)+ ' is an invalid id')

class UserAppInfoApiview(APIView):
    def get(self, request):
        user_email = request.user
        print("--------> Current User", user_email)
        current_user = User.objects.filter(email=user_email)
        if len(current_user) > 0:
            if current_user[0].role == "Auditor":
                app_info = AppInfo.objects.filter(auditor_id=current_user[0].id)
            if current_user[0].role == "Reviewer":
                app_info = AppInfo.objects.filter(reviewer_id=current_user[0].id)
            serializer = AppInfoSerializer(app_info, many=True)
            return Response({'status-code':200, 'errors':[], 'payload':serializer.data},status=status.HTTP_200_OK)
        return Response({'status-code':500, 'message':'User does not exists'})
    
    def patch(self, request):
        current_user = User.objects.filter(email=request.user)
        app_info_data = request.data
        aud_id = app_info_data.get('auditor_id')
        existing_app = AppInfo.objects.filter(app_name=app_info_data['app_name'])   
        if len(existing_app) > 0:
            existing_aud = existing_app[0].auditor_id
            existing_rev = existing_app[0].reviewer_id
            if aud_id != existing_aud:
                app_info_data['auditor_assigned_date'] = date.today() # This logic is for if new reviewer is assigned
            serializer = AppInfoSerializer(existing_app[0], data = app_info_data, partial=True) 
            if serializer.is_valid():
                serializer.save()
                return Response({'status-code':200, 'errors':[], 'message':'Data updated Successfully..!'})
            return Response({'status-code':500, 'message':'Invalid Data'})
        else:
            raise serializers.ValidationError('Applicationsadasd Does not exists..!')

class ExcelFileUploadView(APIView):
    def is_valid_excel(self, file_obj):
        try:
            # Attempt to load the file as an Excel workbook
            file_obj.seek(0)
            wb = load_workbook(file_obj, read_only=True)
            # Check if the workbook contains any sheets
            if not wb.sheetnames:
                print("2nd")
                return False
            # Check if the first sheet has any rows
            first_sheet = wb[wb.sheetnames[0]]
            print("first_sheet", first_sheet)
            if first_sheet.max_row == 0:
                print("3rd")
                return False
            # If all checks pass, return True
            return True
        except Exception:
            return False

    parser_classes = [MultiPartParser]
    def post(self, request):
        file_obj = request.FILES['file']
        request_data = request.data

        # Check file type and content
        if not self.is_valid_excel(file_obj):
            return Response({'status_code': 400, 'errors': ['Unable to read the file'], 'message': 'Only Excel files (.xlsx) are allowed'},
                            status=status.HTTP_400_BAD_REQUEST)

        try:
            # Read the uploaded Excel file
            df = pd.read_excel(file_obj)
            checklist_id = request_data['checklist_id']
            checklist_id = int(checklist_id)
            questions_data = []



            if 'category' not in df.columns or 'question' not in df.columns:
                raise serializers.ValidationError({'message': 'Excel file must contain "category" and "question" columns'})

            checklist = ChecklistData.objects.filter(id=checklist_id).first()

            # Validate if checklist exists
            if not checklist:
                raise serializers.ValidationError({'message': f'Checklist with ID {checklist_id} does not exist'})

            categories = set(df['category'].apply(lambda x: str(x).capitalize()))
            # if categories contain Not a number then remove it from the set
            categories = {x for x in categories if x != 'Nan'}
            print("categories", categories)

            # Validate if all categories in Excel exist in checklist
            missing_categories = categories - set(checklist.subcategories)
            if missing_categories:
                raise serializers.ValidationError({'message': f'Categories {", ".join(missing_categories)} do not exist in checklist'})


             # Check if questions already exist for categories in the database
            existing_questions = QuestionData.objects.filter(checklist_id=checklist_id)
            existing_data = existing_questions.values_list('subcategory', 'question_text')
            existing_set = {(subcategory.capitalize(), question_text.capitalize()) for subcategory, question_text in existing_data}
            
            new_data = {(str(row.get('category')).capitalize(), str(row.get('question')).capitalize()) for index, row in df.iterrows()}
            
            if existing_set.intersection(new_data):
                raise serializers.ValidationError({'message': 'Some questions that are present in excel file already exist in the database.'})

            for index, row in df.iterrows():
                category = str(row.get('category')).capitalize()
                print(category)
                question = str(row.get('question')).capitalize()
                print(question)

                # Validate if the required fields are present
                if not all([category, question]):
                    return Response({'status_code': 400, 'errors': [],
                                     'message': 'Missing required fields in the Excel file'},
                                    status=status.HTTP_400_BAD_REQUEST)

                # Check if the question already exists
                # existing_question = QuestionData.objects.filter(checklist_id=checklist_id,
                #                                                  subcategory=category,
                #                                                  question_text=question)


                # if existing_question.exists():
                #     return Response({'status_code': 400, 'errors': [],
                #                      'message': f'{question} already exists under subcategory '
                #                                 f'{str(category)} in checklist_id {str(checklist_id)}'},
                #                     status=status.HTTP_400_BAD_REQUEST)

                questions_data.append({
                    'checklist_id': checklist_id,
                    'subcategory': category,
                    'question_text': question
                })

                # Remove the objects where either subcategory or question is empty or None or Nan
                questions_data = [qd for qd in questions_data if qd['subcategory'] and qd['question_text'] and qd['subcategory'] != 'Nan' and qd['question_text'] != 'Nan']
                print("questions_data", questions_data)       

            # Bulk create questions
            serializer = QuestionDataSerializer(data=questions_data, many=True)
            
            print("-----------this is questions data qd",questions_data)
            if serializer.is_valid():
                print("qd",questions_data)
                serializer.save(checklist_id=checklist_id)
                return Response({'status_code': 201, 'errors': [], 'message': 'Questions created successfully'},
                                status=status.HTTP_200_OK)
            else:
                return Response({'status_code': 400, 'errors': serializer.errors, 'message': 'Invalid Data'},
                                status=status.HTTP_400_BAD_REQUEST)

            # if serializer.is_valid():
            #     questions_data = QuestionData.objects.create(
            #         checklist_id =checklist_id,
            #         question_text = question,
            #         subcategory = category
            #     )
            #     if question_data:
            #         print('This is question details--------->', question_data)
            #         return Response({'message':'created succeffully'})

        except serializers.ValidationError as ve:
            return Response({'status_code': 400, 'errors': [ve.detail['message']], 'message': ''},
                            status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'status_code': 500, 'errors': [str(e)], 'message': 'Error processing Excel file'},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class ChecklistDataCreateView(APIView):
    def get(self,request):
        checklist_data = ChecklistData.objects.all()
        serializer = ChecklistDataSerializer(checklist_data, many=True)
        return Response({'status_code':200, 'errors': [], 'payload':serializer.data},status=status.HTTP_200_OK)

    def post(self,request):
        checklist_title = str(request.data.get('checklist_title')).capitalize()
        subcategories = set()
        for cat in request.data.get('subcategories'):
            subcategories.add(str(cat).capitalize())
        data = dict(checklist_title=checklist_title, subcategories=list(subcategories))
        existing_checklist = ChecklistData.objects.filter(checklist_title=checklist_title)
        if len(existing_checklist) > 0:
            return Response({'status_code':400, 'errors':[], 'message':checklist_title+' already exists'}, status=status.HTTP_400_BAD_REQUEST)
        serializer = ChecklistDataSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response({'status_code':201, 'errors':[], 'message':'Checklist Created successfully'},status=status.HTTP_201_CREATED)
        return Response({'status_code':400, 'errors':serializer.errors,'message':'Invalid Data'},status=status.HTTP_400_BAD_REQUEST)   


class ChecklistDataDetailView(APIView):
    def get_checklist(self, id):
        try:
            return ChecklistData.objects.get(id=id)
        except ChecklistData.DoesNotExist():
            raise status.HTTP_404_NOT_FOUND

    def get(self, request, id):
        try:
            checklist = self.get_checklist(id)
            serializer = ChecklistDataSerializer(checklist)
            return Response({'status_code':200, 'errors':[], 'payload':serializer.data},status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'status_code':400, 'errors':[], 'message': "checklist with "+str(id)+" does not exists."},status=status.HTTP_400_BAD_REQUEST)
            
    def patch(self,request, id):
        try:
            checklist = self.get_checklist(id)
            checklist_title = str(request.data.get('checklist_title')).capitalize()
            # all_subcategories = [*checklist.subcategories, *request.data.get('subcategories')]
            all_subcategories = [*request.data.get('subcategories')]
            subcategories = set()
            for cat in all_subcategories:
                subcategories.add(str(cat).capitalize())
            
            data = dict(checklist_title=checklist_title, subcategories=list(subcategories))
            serializer = ChecklistDataSerializer(checklist, data=data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response({'status_code':200, 'errors':[], 'message':'Checklist updated successfully'}, status=status.HTTP_200_OK)
            return Response({'status_code':400, 'errors':serializer.errors, 'message': "checklist "+checklist_title+" does not exists."},status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'status_code':400, 'errors':[], 'message': "checklist with "+str(id)+" does not exists."},status=status.HTTP_400_BAD_REQUEST)
       
    def delete(self,request,id):
        try:
            checklist = self.get_checklist(id)
            checklist.delete()
            return Response({'status_code':204, 'errors':[], 'message':'Data Deleted successfully'}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'status_code':400, 'errors':[], 'message': "checklist with "+str(id)+" does not exists."},status=status.HTTP_400_BAD_REQUEST)
    
class QuestionDataCreateView(APIView):
    def get(self, request):
        try:
            question_data = QuestionData.objects.all()
            serializer = QuestionDataSerializer(question_data, many=True)
            return Response({'status_code':200, 'errors': [], 'payload':serializer.data},status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'status_code':400, 'errors':[], 'message': "Question with "+str(id)+" does not exists."},status=status.HTTP_400_BAD_REQUEST)

    def post(self, request):
        checklist_id = request.data.get('checklist_id')
        subcategory = str(request.data.get('subcategory')).capitalize()
        checklist = ChecklistData.objects.filter(id=checklist_id)
        question_text = str(request.data.get('question_text')).capitalize()
        data = dict(checklist_id=checklist_id, subcategory=subcategory, question_text=question_text)

        print("-----Befor---->",data)
        existing_question = QuestionData.objects.filter(checklist_id=checklist_id,subcategory=subcategory,question_text=question_text)
        print("-----existing_question---->",existing_question)
        if len(existing_question) > 0:  
            return Response({'status_code':400, 'errors':[], 'message':question_text+' already exists under subategory '+subcategory+' in checklist_id '+ str(checklist_id)}, status=status.HTTP_400_BAD_REQUEST)   

        if len(checklist) > 0:
            if subcategory not in checklist[0].subcategories:
                return Response({'status_code': 400, 'errors':[],'message': subcategory + ' does not exists in checklist data'}, status=status.HTTP_400_BAD_REQUEST)
            print("-----After---->",data)
            serializer = QuestionDataSerializer(data=data)
            if serializer.is_valid():
                # checklist_id = serializer.validated_data.get('checklist_id')
                subcategory = serializer.validated_data.get('subcategory')
                question_text = serializer.validated_data.get('question_text')
                print("-----After---ser->",data)
                try:
                    question = QuestionData.objects.create(
                        checklist_id=checklist_id,
                        subcategory=subcategory,
                        question_text=question_text
                    )
                    if question: 
                        print("quws->",question)
                        return Response({'status_code':201, 'errors':[], 'message':'Question Created successfully'},status=status.HTTP_200_OK)
                except Exception as e:
                    return Response({'status_code':400, 'errors':[],'message':'This Question already exists in this Checklist under another subcategory'},status=status.HTTP_400_BAD_REQUEST)
                return Response({'status_code':400, 'errors':serializer.errors,'message':'Invalid Data'},status=status.HTTP_400_BAD_REQUEST)   
        return Response({'status_code':404, 'errors':[], 'message':'Checklist with '+str(checklist_id)+ ' does not exists'},status=status.HTTP_400_BAD_REQUEST)      

class QuestionDataDetailView(APIView):
    def get_question(self, id):
        try:
            return QuestionData.objects.get(id=id)
        except QuestionData.DoesNotExist():
            raise status.HTTP_404_NOT_FOUND
    
    def get(self, request, id):
        try:
            question = self.get_question(id)
            serializer = QuestionDataSerializer(question)
            return Response({'status_code':200, 'errors':[], 'payload':serializer.data},status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'status_code':400, 'errors':[], 'message': "Question with "+str(id)+" does not exists."},status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request, id):
        try:
            question = self.get_question(id)
            checklist_id = request.data.get('checklist_id')
            subcategory = str(request.data.get('subcategory')).capitalize()
            checklist = ChecklistData.objects.filter(id=checklist_id)
            question_text = str(request.data.get('question_text')).capitalize()
            data = dict(checklist_id=checklist_id, subcategory=subcategory, question_text=question_text)
            if len(checklist) > 0:
                if subcategory not in checklist[0].subcategories:
                    return Response({'status_code': 400, 'errors':[],'message': subcategory + ' does not exists in checklist data'}, status=status.HTTP_400_BAD_REQUEST)
                serializer = QuestionDataSerializer(question, data=data, partial=True)
                if serializer.is_valid():
                    serializer.save()
                    return Response({'status_code':200, 'errors':[], 'message':'Question updated successfully'}, status=status.HTTP_200_OK)
                return Response({'status_code':400, 'errors':serializer.errors},status=status.HTTP_400_BAD_REQUEST)
            return Response({'status_code':404, 'errors':[], 'message':'Checklist with '+str(checklist_id)+ ' does not exists'},status=status.HTTP_400_BAD_REQUEST)   
        except Exception as e:
            return Response({'status_code':400, 'errors':[], 'message': "Question with "+str(id)+" does not exists."},status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, id):
        try:
            question = self.get_question(id)
            question.delete()
            return Response({'status_code':204, 'errors':[], 'message':'Question Deleted successfully'}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'status_code':400, 'errors':[], 'message': "Question with "+str(id)+" does not exists."},status=status.HTTP_400_BAD_REQUEST)

class AppInfoViewForAdmin(APIView):
    
    def post(self,request): #Create and assign application to reviewer
        app_info_data = request.data
        if not request.data.get('app_name'):
            raise serializers.ValidationError("Application name is required !")
        existing_app = AppInfo.objects.filter(app_name=app_info_data['app_name'],checklist_id=app_info_data['checklist_id'])
        if len(existing_app) > 0: # chk if app already exists in db
            return Response({'status_code':400, 'message': app_info_data['app_name'] + ' already exists ..!'},status=status.HTTP_400_BAD_REQUEST)
        rev_id = app_info_data.get('reviewer_id') # get reviewer from request
        reviewer = User.objects.filter(id=rev_id) # chk if the reviewer exists in db
        if len(reviewer)>0:
            if reviewer[0].role not in ['Reviewer' , 'reviewer']:
                raise serializers.ValidationError(str(rev_id) + " is not a Reviewer!")
        else:
            raise serializers.ValidationError("Reviewer Does Not exists with id:" + str(rev_id))
        if rev_id is not None: # if rev_id is present in request the set reviewer assign date as today
            user_email = request.user
            current_user = User.objects.filter(email=user_email)[0] # getting current user by its email
            action = "[{}] {} ({})".format(current_user.role, current_user.full_name, user_email)
            if current_user.role == "Admin":
                print("This is current user--****->",current_user.role)
                todayDate = date.today().isoformat()
                app_info_data['reviewer_assigned_date'] = todayDate
                app_info_data['timeline'] = [{'status_explanation':'Application assigned to Reviewer '+reviewer[0].full_name, 
                                                    'date': todayDate, 'action_by': action}]
            else:
                return Response({'status_code':400, 'errors':'[]', 'message':current_user.full_name+ ' is not an Admin'},status=status.HTTP_400_BAD_REQUEST)
        serializer = AppInfoSerializer(data=app_info_data)
        if serializer.is_valid():
            serializer.save()
            return Response({'status_code':200, 'error':[],'message':'Application added successfully','payload':serializer.data},status=status.HTTP_200_OK)
        return Response({'status_code':400, 'message':'Invalid Option', 'errors': serializer.errors},status=status.HTTP_400_BAD_REQUEST)
    
    def get(self, request):
        app_info = AppInfo.objects.all()
        serializer = AppInfoSerializer(app_info, many=True)
        return Response({'status_code':200, 'payload':serializer.data},status=status.HTTP_200_OK)

    def patch(self, request, id): # update app_info and change reviewer
        app_info_data = request.data
        rev_id = app_info_data.get('reviewer_id')
        existing_app = AppInfo.objects.filter(id=id) # getting existing app from db
        if len(existing_app) > 0:
            existing_rev = existing_app[0].reviewer_id 
            if rev_id is not None:
                if rev_id != existing_rev: # if request rev_id is not equal to already assigned revid then only update it
                    user = request.user
                    current_user = User.objects.filter(email=user)[0]
                    reviewer = User.objects.filter(id=rev_id)
                    action = "[{}] {} ({})".format(current_user.role, current_user.full_name, user)
                    if current_user.role == "Admin":
                        todayDate = date.today().isoformat()
                        app_info_data['reviewer_assigned_date'] = todayDate
                        timeline_data = {'status_explanation':'Application assigned to Reviewer '+ reviewer[0].full_name,'date': todayDate, 'action_by': action}
                        app_info_data['timeline'] = existing_app[0].timeline    
                        app_info_data['timeline'].append(timeline_data)
                        app_info_data['status'] = "assigned_to_rev"
                    else:
                        return Response({'status_code':400, 'errors':'[]', 'message':current_user.full_name+ ' is not an Admin'},status=status.HTTP_400_BAD_REQUEST)
                serializer = AppInfoSerializer(existing_app[0], data = app_info_data, partial=True) 
                if serializer.is_valid():
                    try:
                        serializer.save()
                    except:
                        return Response({'status_code':400, 'errors':[], 'message':'something went wrong while saving data!'}, status=status.HTTP_400_BAD_REQUEST)
                    return Response({'status_code':200, 'errors':[], 'message':'Data updated Successfully..!'}, status=status.HTTP_200_OK)
                return Response({'status_code':400, 'errors':serializer.errors, 'message':''}, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'status_code':400, 'errors':[], 'message':'Application does not exists.!'}, status=status.HTTP_400_BAD_REQUEST)

    def delete(self,request, id):
        # app_id = request.data.get('id')
        try:
            app_data = AppInfo.objects.get(id=id)
            app_data.delete()
            return Response({'status_code':200, 'errors':[],'message':'Application deleted successfully'})
        except:
            raise serializers.ValidationError(str(id)+ ' is an invalid id')

class AppInfoViewForReviewer(APIView):
    def patch(self, request,id):
        app_info_data = request.data
        existing_app = AppInfo.objects.filter(id=id)
        if len(existing_app) > 0:
            aud_id = app_info_data.get('auditor_id')
            auditor = User.objects.filter(id=aud_id)
            if len(auditor) > 0:
                if auditor[0].role not in ['Auditor', 'auditor']:
                    raise serializers.ValidationError(str(aud_id) + " is not an Auditor!")
            else:
                raise serializers.ValidationError("Auditor Does Not exists with id:" + str(aud_id))
            if aud_id is not None:
                if aud_id == existing_app[0].auditor_id:
                    return Response({'status_code':400, 'errors':'[]', 'message':existing_app[0].app_name+ ' is already assigned to '+ str(aud_id)},status=status.HTTP_400_BAD_REQUEST)
                user_email = request.user
                current_user = User.objects.filter(email=user_email)[0]
                action = "[{}] {} ({})".format(current_user.role, current_user.full_name, user_email)
                todayDate = date.today().isoformat()
                app_info_data['auditor_assigned_date'] = todayDate
                app_info_data['timeline'] = existing_app[0].timeline
                timeline_data = {'status_explanation':'Application assigned to Auditor '+auditor[0].full_name, 'date': todayDate, 'action_by': action}
                app_info_data['timeline'].append(timeline_data)
                app_info_data['status'] = "assigned_to_aud"
                print('app--->',app_info_data)
            serializer = AppInfoSerializer(existing_app[0],data=app_info_data,partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response({'status_code':200, 'error':[],'message':'Application assigned successfully','payload':serializer.data},status=status.HTTP_200_OK)
            return Response({'status_code':400, 'message':'Invalid Option', 'errors': serializer.errors},status=status.HTTP_400_BAD_REQUEST)
        return Response({'status_code':400, 'message':'Application does not exist!', 'errors': []},status=status.HTTP_400_BAD_REQUEST)

 
class GetAppInfoView(APIView):
    def get(self, request):
        user_email = request.user
        current_user = User.objects.filter(email=user_email).first()
        apps = []

        if user_email is not None:
            if current_user.role == 'Reviewer':
                apps = AppInfo.objects.filter(reviewer_id=current_user.id)
            elif current_user.role == 'Auditor':
                apps = AppInfo.objects.filter(auditor_id=current_user.id)
        else:
            apps = AppInfo.objects.all()

        serializer = AppInfoSerializer(apps, many=True)
        app_info_data = []
        for app_data in serializer.data:
            reviewer = User.objects.filter(id=app_data['reviewer_id']).first()
            auditor = User.objects.filter(id=app_data['auditor_id']).first()

            reviewer_name = reviewer.full_name if reviewer else None
            auditor_name = auditor.full_name if auditor else None

            data = {
                **app_data,
                'reviewer_name': reviewer_name,
                'auditor_name': auditor_name
            }
            app_info_data.append(data)

        return Response({'status_code': 200, 'errors': [], 'payload': app_info_data}, status=status.HTTP_200_OK)

class AnswerDataView(APIView):
    def post(self, request):
        app_detail_id = request.data.get('app_detail_id')

        existing_app_detail = AppDetail.objects.filter(id=app_detail_id)
        if len(existing_app_detail) == 0:
            return Response({'status_code':400, 'message':'Invalid app_detail_id: '+str(app_detail_id), 'errors':[]},status=status.HTTP_400_BAD_REQUEST)

        subcategory = str(request.data.get('subcategory')).capitalize()
        question_answer = request.data.get('question_answer') # list of question_answer
        for item in question_answer:
            # consider POC upload using pillow
            # POC = item    ?????
            question = str(item['question_text']).capitalize()
            answer = str(item['answer_text']).capitalize()

            # validation to see if question is already answered
            existing_answer = AnswerData.objects.filter(app_detail_id=app_detail_id,subcategory=subcategory, question_text=question, answer_text=answer)
            if len(existing_answer) > 0:
                return Response({'status_code':400, 'message':'Answer Data is already saved for '+ question, 'errors': []},status=status.HTTP_400_BAD_REQUEST)
            data = dict(app_detail_id=app_detail_id, subcategory=subcategory, question_text=question, answer_text=answer)  
            serializer = AnswerDataSerializer(data=data)
            if serializer.is_valid():
                serializer.save()
            else:
                return Response({'status_code':400, 'message':'Invalid data', 'errors': serializer.errors},status=status.HTTP_400_BAD_REQUEST)
        return Response({'status_code':200, 'error':[],'message':'Question Answer saved successfully'},status=status.HTTP_200_OK)

    def patch(self, request):
        question_answer = request.data.get('question_answer', [])
        for item in question_answer:
            app_detail_id = item.get('app_detail_id')
            subcategory = str(item.get('subcategory', '')).capitalize()
            question_text = str(item.get('question_text', '')).capitalize()
            answer_text = str(item.get('answer_text', '')).capitalize()

            if not app_detail_id:
                return Response({'status_code': 400, 'message': 'app_detail_id is required', 'errors': []}, status=status.HTTP_400_BAD_REQUEST)

            # Get the existing answer if it exists
            existing_answer = AnswerData.objects.filter(app_detail_id=app_detail_id, subcategory=subcategory, question_text=question_text).first()

            if existing_answer:
                # If an existing entry is found, update it
                existing_answer.answer_text = answer_text
                existing_answer.save()
            else:
                # If no existing entry is found, create a new one
                data = {'app_detail_id': app_detail_id, 'subcategory': subcategory, 'question_text': question_text, 'answer_text': answer_text}
                serializer = AnswerDataSerializer(data=data)

                if serializer.is_valid():
                    serializer.save()
                else:
                    return Response({'status_code': 400, 'message': 'Invalid data', 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

        return Response({'status_code': 200, 'error': [], 'message': 'Question Answer Updated/Added successfully'}, status=status.HTTP_200_OK)
    
class AppDetailView(APIView):
    def create_app_feilds(self, id, key, value):
        existingkey = AppField.objects.filter(key=key, app_detail_id=id)
        if len(existingkey) > 0:
            return Response({'status_code':400, 'errors':[], 'message': key +' already Exists.!'}, status=status.HTTP_400_BAD_REQUEST)
        
        data = dict(app_detail_id=id, key=key, value=value)
        serializer = AppFieldSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response({'status_code': 200, 'errors': [], 'message': 'App Field created successfully'}, status=status.HTTP_200_OK)
        else:
            return Response({'status_code':400, 'errors':serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    def post(self, request): # Create Key Value for app_field and subcategory weightage
        user_email = request.user
        current_user = User.objects.get(email=user_email)
        app_info_id = request.data.get('app_info_id')
        unique_field = request.data.get('unique_field')
        subcategory_weightage = request.data.get('subcategory_weightage')
        app_fields = request.data.get('app_fields')

        # validation for AppInfo Id exists
        app_info = AppInfo.objects.filter(id=app_info_id)
        if len(app_info) == 0:
            return Response({'status_code': 400, 'message': 'Invalid app_info_id provided.'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Validation for Subcategory
        checklist = ChecklistData.objects.filter(id=app_info[0].checklist_id)[0]
        for key,value in subcategory_weightage.items():
            key = str(key).capitalize()
            if key not in checklist.subcategories:
                return Response({'status_code':400, 'message':key +' Subcategory not present in checklist.'},status=status.HTTP_400_BAD_REQUEST)
            if not (0 <= value <=5 ) :
                return Response({'status_code':400, 'message':'Subcategory weightage must be between 0 to 5.'},status=status.HTTP_400_BAD_REQUEST)

        # validation on unique_field
        # Check if unique_field value is present in app_fields
        unique_field_present = False
        for field in app_fields:
            if field["key"] == unique_field:
                unique_field_present = True
                break

        if unique_field_present == False:
            return Response({'status_code':400, 'message':'Unique Field is not present in App Fields.'},status=status.HTTP_400_BAD_REQUEST)

        # To avoid Multiple details of same application
        existing_app_detail = AppDetail.objects.filter(app_info_id=app_info_id)
        if len(existing_app_detail) > 0:
            return Response({'status_code':400, 'message':'App Details already exists for this application'},status=status.HTTP_400_BAD_REQUEST)

        data = dict(app_info_id=app_info_id, unique_field=unique_field, subcategory_weightage=subcategory_weightage)
        serializer = AppDetailSerializer(data=data)
        if serializer.is_valid():
            app_info_id = serializer.validated_data.get('app_info_id')
            unique_field = serializer.validated_data.get('unique_field')   
            subcategory_weightage = serializer.validated_data.get('subcategory_weightage')   
            created = AppDetail.objects.create(
                    app_info_id=app_info_id,
                    unique_field=unique_field,
                    subcategory_weightage=subcategory_weightage
            )
            print("App Details--->", created.id)
            if created:
                for field in app_fields:
                    response = self.create_app_feilds(created.id, field['key'], field['value'])
                    # Check if response indicates an error, and if so, return it immediately
                    if response.status_code != status.HTTP_200_OK:
                        return response
                application = AppInfo.objects.get(id=app_info_id)
                todayDate = date.today().isoformat()
                action = "[{}] {} ({})".format(current_user.role, current_user.full_name, user_email)
                application.status = "inprogress"
                application.timeline = [{'status_explanation':'Application is in Progress.', 
                                                    'date': todayDate, 'action_by': action}]
                application.save()
                return Response({'status_code':200, 'errors':[],'message':'App Detail Added successfully'}, status=status.HTTP_200_OK)
            else:
                return Response({'status_code': 400, 'errors': serializer.errors, 'message': 'Validation errors occurred.'}, status=status.HTTP_400_BAD_REQUEST)

    def get(self,request, app_info_id):# get by app_info_id
        app_detail = AppDetail.objects.filter(app_info_id=app_info_id)
        if len(app_detail) > 0:
            serializer = AppDetailSerializer(app_detail[0])
            return Response({'status_code':200, 'payload': serializer.data}, status=status.HTTP_200_OK)
        return Response({'status_code': 400, 'errors': [], 'message': 'Invalid App Info Id.'}, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, app_info_id): # Delete app info by app_info_id
        app_detail = AppDetail.objects.filter(app_info_id=app_info_id)
        if len(app_detail) > 0:
            app_detail[0].delete()
            return Response({'status_code':200, 'message':'App Detail Deleted Successfully'}, status=status.HTTP_200_OK)
        return Response({'status_code': 400, 'errors': [], 'message': 'Invalid App Info Id.'}, status=status.HTTP_400_BAD_REQUEST)

    def patch(self,request,app_info_id): # To edit unique_feild, subcategory_weightage, app_feild or create new app_feilds
        unique_field = request.data.get('unique_field')
        subcategory_weightage = request.data.get('subcategory_weightage')

        app_info = AppInfo.objects.filter(id=app_info_id)

        if len(app_info) == 0:
            return Response({'status_code': 400, 'message': 'Invalid app_info_id provided.'}, status=status.HTTP_400_BAD_REQUEST)

        checklist = ChecklistData.objects.filter(id=app_info[0].checklist_id)[0]
        for key,value in subcategory_weightage.items():
            key = str(key).capitalize()
            if key not in checklist.subcategories:
                return Response({'status_code':400, 'message':str(key) + ' Subcategory does not exists in Checklist'},status=status.HTTP_400_BAD_REQUEST)
            if not (0 <= value <=5 ) :
                return Response({'status_code':400, 'message':'Subcategory weightage must be between 0 to 5.'},status=status.HTTP_400_BAD_REQUEST)

        app_detail = AppDetail.objects.filter(app_info_id=app_info_id)
        existing_unique_field = AppField.objects.filter(app_detail_id=app_detail[0].id, key=unique_field)
        if len(existing_unique_field) == 0:
            return Response({'status_code':400, 'message':'Unique Field is not present in App Fields.'},status=status.HTTP_400_BAD_REQUEST)
        print('sub--->',subcategory_weightage)

        app_feild_operation_error = []
        
        # Checking if app_field from request exists then only we can add or edit key value:
        if 'app_fields' in request.data and len(request.data.get('app_fields')) > 0: # check app_fields exists in request data 
            app_fields = request.data.get('app_fields')
            print('paythc--->', app_fields)
            for item in app_fields:
                existing_field = AppField.objects.filter(app_detail_id=app_detail[0].id,key=item['key']) 
                if len(existing_field) > 0: # edit key value fields if exists
                    app_field_data = dict(key=item['key'], value=item['value'])
                    app_field_serializer = AppFieldSerializer(existing_field[0], data=app_field_data, partial=True)
                    if app_field_serializer.is_valid():
                        app_field_serializer.save()
                    else:
                        app_feild_operation_error.append(app_field_serializer.errors) 
                else: # used utility function to create key value fields, if error utility function will return error object and append it
                    app_feild_operation_error.append(self.create_app_feilds(app_detail[0].id, item['key'], item['value']) )
        
        data = dict(unique_field=unique_field, subcategory_weightage=subcategory_weightage)
        serializer = AppDetailSerializer(app_detail[0], data=data, partial=True)
        if serializer.is_valid():
           serializer.save()
           errors = []
           message = 'App Detail Edited successfully'
           print('serializwerappdeat--->',app_feild_operation_error)
        #    if len(app_feild_operation_error) > 0:
        #        errors = app_feild_operation_error
        #        message = message + ' but encountered error while app feild operations'
           return Response({'status_code':200, 'errors': errors,'message': message}, status=status.HTTP_200_OK)
        else:
            return Response({'status_code': 400, 'errors': serializer.errors, 'message': 'Validation errors occurred.'}, status=status.HTTP_400_BAD_REQUEST)

class AppFieldView(APIView):
    def patch(self, request, app_field_id):
        key = request.data.get('key')
        value = request.data.get('value')
        data = dict(key=key, value=value)
        app_field = AppField.objects.filter(id=app_field_id)
        if len(app_field) > 0:
            app_detail = AppDetail.objects.filter(id=app_field[0].app_detail_id)[0]
            if key == app_detail.unique_field:
                return Response({'status_code': 400, 'errors': [], 'message': key+' is a Unique Field so cannot be updated.'}, status=status.HTTP_400_BAD_REQUEST)
            serializer = AppFieldSerializer(app_field[0], data=data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response({'status_code':200, 'errors':[],'message':'App Field Updated successfully'}, status=status.HTTP_200_OK)
            else:
                return Response({'status_code': 400, 'errors': serializer.errors, 'message': 'Validation errors occurred.'}, status=status.HTTP_400_BAD_REQUEST)
        return Response({'status_code': 400, 'errors': [], 'message': 'App Field id does not exists.'}, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, app_field_id):
        app_field = AppField.objects.filter(id=app_field_id)
        print("==============================================================================app_field", app_field)
        if len(app_field) > 0:
            app_detail = AppDetail.objects.filter(id=app_field[0].app_detail_id)[0]
            if app_field[0].key == app_detail.unique_field:
                return Response({'status_code': 400, 'errors': [], 'message': app_field[0].key+' is a Unique Field so cannot be Deleted.'}, status=status.HTTP_400_BAD_REQUEST)
            app_field[0].delete()
            return Response({'status_code':200, 'message':'App Field Deleted Successfully'}, status=status.HTTP_200_OK)
        return Response({'status_code': 400, 'errors': [], 'message': 'Invalid App Field Id.'}, status=status.HTTP_400_BAD_REQUEST)
    
class AppCountView(APIView):
    def get(self, request):
        user_email = request.user
        current_user = User.objects.get(email=user_email)
        response = dict()
        if current_user.role == 'Admin':
            assigned_to_rev = AppInfo.objects.filter(status='assigned_to_rev').count()
            assigned_to_aud = AppInfo.objects.filter(status='assigned_to_aud').count()
            submitted_to_rev = AppInfo.objects.filter(status='submitted_to_rev').count()
            submitted_to_admin = AppInfo.objects.filter(status='submitted_to_admin').count()
            total_apps = AppInfo.objects.all().count()
            completed_apps = AppInfo.objects.filter(status='completed').count()
            inprogress_apps = AppInfo.objects.filter(status='inprogress').count()
            onhold_apps = AppInfo.objects.filter(status='hold').count()
            total_reviewers = User.objects.filter(role='Reviewer').count()
            total_auditors = User.objects.filter(role='Auditor').count()
            response = {'assigned_to_rev':assigned_to_rev, 'assigned_to_aud':assigned_to_aud,
                'submitted_to_rev':submitted_to_rev,
                'submitted_to_admin':submitted_to_admin, 'onhold_apps':onhold_apps,
                'total_apps': total_apps, 'completed_apps': completed_apps,
                'inprogress_apps': inprogress_apps, 'total_reviewers': total_reviewers,
                'total_auditors':total_auditors}
        elif current_user.role == 'Reviewer':
            total_apps = AppInfo.objects.filter(reviewer_id = current_user.id).count()
            completed_apps = AppInfo.objects.filter(status='completed',reviewer_id = current_user.id).count()
            assigned_to_rev = AppInfo.objects.filter(status='assigned_to_rev',reviewer_id = current_user.id).count()
            assigned_to_aud = AppInfo.objects.filter(status='assigned_to_aud',reviewer_id = current_user.id).count()
            submitted_to_rev = AppInfo.objects.filter(status='submitted_to_rev',reviewer_id = current_user.id).count()
            submitted_to_admin = AppInfo.objects.filter(status='submitted_to_admin',reviewer_id = current_user.id).count()
            inprogress_apps = AppInfo.objects.filter(status='inprogress',reviewer_id = current_user.id).count()
            onhold_apps = AppInfo.objects.filter(status='hold',reviewer_id = current_user.id).count()
            response = {'assigned_to_rev':assigned_to_rev, 'assigned_to_aud':assigned_to_aud,
                'submitted_to_rev':submitted_to_rev,
                'submitted_to_admin':submitted_to_admin, 'onhold_apps':onhold_apps,
                'total_apps': total_apps, 'completed_apps': completed_apps,
                'inprogress_apps': inprogress_apps}
        elif current_user.role == 'Auditor':
            total_apps = AppInfo.objects.filter(auditor_id = current_user.id).count()
            completed_apps = AppInfo.objects.filter(status='completed',auditor_id = current_user.id).count()
            assigned_to_aud = AppInfo.objects.filter(status='assigned_to_aud',auditor_id = current_user.id).count()
            submitted_to_rev = AppInfo.objects.filter(status='submitted_to_rev',auditor_id = current_user.id).count()
            submitted_to_admin = AppInfo.objects.filter(status='submitted_to_admin',auditor_id = current_user.id).count()
            inprogress_apps = AppInfo.objects.filter(status='inprogress',auditor_id = current_user.id).count()
            onhold_apps = AppInfo.objects.filter(status='hold',auditor_id = current_user.id).count()
            response = {'assigned_to_aud':assigned_to_aud, 'assigned_to_aud':assigned_to_aud,
                'submitted_to_rev':submitted_to_rev,
                'submitted_to_admin':submitted_to_admin, 'onhold_apps':onhold_apps,
                'total_apps': total_apps, 'completed_apps': completed_apps,
                'inprogress_apps': inprogress_apps}
        return Response(response, status=status.HTTP_200_OK)

class StatusChangeView(APIView):
    def getAppSerializer(self, status_explanation,app_info,action,app_status):
        todayDate = date.today().isoformat()
        timeline_data = {'status_explanation': status_explanation, 'date': todayDate, 'action_by': action}
        timeline = app_info[0].timeline
        timeline.append(timeline_data)
        data = dict(status=app_status, timeline=timeline) 
        serializer = AppInfoSerializer(app_info[0],data=data,partial=True)
        return serializer

    def patch(self, request):
        user_email = request.user
        current_user = User.objects.get(email=user_email)
        app_id = request.data.get('app_id')
        app_status = request.data.get('status')
        app_info = AppInfo.objects.filter(id=app_id)
        status_explanation = ''
        action = "[{}] {} ({})".format(current_user.role, current_user.full_name, user_email)

        if current_user.role == 'Admin':
            if len(app_info) > 0:
                if app_status == 'submitted_to_rev':
                    status_explanation = 'Application submitted to Reviewer.'
                # elif app_status == 'submitted_to_admin':
                #     status_explanation = 'Application submitted to Admin.'
                elif app_status == 'hold':
                    status_explanation = 'Application on hold.'
                elif app_status == 'completed':
                    status_explanation = 'Application assessment completed.'
                serializer = self.getAppSerializer(status_explanation, app_info, action, app_status)
                if serializer.is_valid():
                    serializer.save()
                    return Response({'status_code':200, 'error':[],'message':'Status updated successfully'},status=status.HTTP_200_OK)
                else:
                    return Response({'status_code':404, 'errors':serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
            else:
                print("error")
                return Response({'status_code':404, 'message':'Invalid App Id'}, status=status.HTTP_400_BAD_REQUEST)
        elif current_user.role == 'Reviewer':
            if len(app_info) > 0:
                if app_status == 'submitted_to_admin':
                    status_explanation = 'Application submitted to Admin.'
                elif app_status == 'completed':
                    status_explanation = 'Application assessment completed.'
                serializer = self.getAppSerializer(status_explanation, app_info, action, app_status)
                if serializer.is_valid():
                    serializer.save()
                    return Response({'status_code':200, 'error':[],'message':'Status updated successfully'},status=status.HTTP_200_OK)
                else:
                    return Response({'status_code':404, 'errors':serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
            else:
                print("error")
                return Response({'status_code':404, 'message':'Invalid App Id'}, status=status.HTTP_400_BAD_REQUEST)
        elif current_user.role == 'Auditor':
            app_detail = AppDetail.objects.filter(app_info_id=app_id)
            app_info = AppInfo.objects.filter(id=app_id)
            if len(app_detail) > 0:
                checklist_id = app_info[0].checklist_id
                total_questions = QuestionData.objects.filter(checklist_id=checklist_id).count()
                answered_questions = AnswerData.objects.filter(app_detail_id=app_detail[0].id).count()
                print("answered_questions", answered_questions)
                print("total_questions", total_questions)
                if answered_questions != total_questions:
                    return Response({'status_code':400, 'errors':[], 'message':'Please save responses of all the questions before submitting.'}, status=status.HTTP_400_BAD_REQUEST)
            if len(app_info) > 0:
                if app_status == 'submitted_to_rev':
                    status_explanation = 'Application submitted to Reviewer.'
                serializer = self.getAppSerializer(status_explanation, app_info, action, app_status)
                if serializer.is_valid():
                    serializer.save()
                    return Response({'status_code':200, 'error':[],'message':'Status updated successfully'},status=status.HTTP_200_OK)
                else:
                    return Response({'status_code':404, 'errors':serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
            else:
                print("error")
                return Response({'status_code':404, 'message':'Invalid App Id'}, status=status.HTTP_400_BAD_REQUEST)
        
class RemarkView(APIView):
    def patch(self, request):
        user_email = request.user
        current_user = User.objects.get(email=user_email)
        remark = request.data.get('remark')
        app_info_id = request.data.get('app_info_id')
        action = request.data.get('action')
        message = ''
        app_detail = AppDetail.objects.filter(app_info_id=app_info_id)
        if len(app_detail) > 0:
            existing_remark = app_detail[0].remarks
            todayDate = date.today().isoformat()
            action_by = "[{}] {} ({})".format(current_user.role, current_user.full_name, user_email)
            remark_data = {'remark':remark, 'date':todayDate ,'action_by':action_by}
            if action == 'Add':
                existing_remark.append(remark_data)
                message = 'Remark Added Successfully'
            elif action == 'Remove':
                index_of_remark = next((i for i, item in enumerate(existing_remark) if item["remark"] == remark and item["action_by"] ==  action_by), None)
                if index_of_remark is None:
                    return Response({'status_code':404,'error':[],'message':'You cannot perform this action' }, status=status.HTTP_400_BAD_REQUEST)
                else:
                    existing_remark.pop(index_of_remark)
                    message = 'Remark Removed Successfully'
            data = dict(remarks=existing_remark)
            serializer = AppDetailSerializer(app_detail[0], data=data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response({'status_code':200, 'error':[],'message': message},status=status.HTTP_200_OK)
            else:
                return Response({'status_code':404,'error':serializer.errors }, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'status_code':404, 'message':'Invalid App Info Id'}, status=status.HTTP_400_BAD_REQUEST)


class ResultView(APIView):
    def calculate_risk_rating(self, pass_count, fail_count):
        pass_fail_sum = pass_count + fail_count
        if pass_fail_sum == 0:
            return 0.0
        risk_rating = (fail_count / pass_fail_sum) * 10
        risk_rating = round(risk_rating, 3)
        return risk_rating

    def post(self, request):
        options = Options.objects.all()
        app_detail_id = request.data.get("app_detail_id")
        app_detail = AppDetail.objects.filter(id=app_detail_id)
        answer_data = AnswerData.objects.filter(app_detail_id=app_detail_id)
        user_email = request.user
        current_user = User.objects.get(email=user_email)
        option_wise_count = {}
        for opt in options:
            # creating {'Pass': 0, 'Fail': 0}
            option_wise_count[opt.option_text] = 0

        if len(app_detail) > 0 and len(answer_data) > 0:
            subcategory_weightage = app_detail[0].subcategory_weightage
            for ans in answer_data:
                answer_subcategory = ans.subcategory # privacy policy, pgrm
                weightage = subcategory_weightage[answer_subcategory] # out of 5
                answer_option = ans.answer_text # answer like Pass, Fail
                prev_count = option_wise_count[answer_option] # existing count
                option_wise_count[answer_option] = prev_count + weightage # add weightage
                # this for will give o/p {'Pass': 6, 'Fail': 3, 'Not applicable': 3, 'Unable to verify': 3}
            print(' subcategory_weightage', subcategory_weightage.keys())
            
            category_wise_rating = {}
            for key in subcategory_weightage.keys():
                # key are Cryptography, Privacy policy, Platform interaction....
                subcategory_wise_ans = AnswerData.objects.filter(app_detail_id=app_detail_id, subcategory=key)
                sub_pass_count = 0
                sub_fail_count = 0
                for sub_ans in subcategory_wise_ans:
                    weightage = subcategory_weightage[key]
                    if sub_ans.answer_text == "Pass":
                        sub_pass_count = sub_pass_count + weightage
                    elif sub_ans.answer_text == "Fail":
                        sub_fail_count = sub_fail_count + weightage
                category_wise_rating[key] = self.calculate_risk_rating(sub_pass_count, sub_fail_count) * 10 # convert into %
            # this for will give o/p {'Cryptography': 33.3, 'Privacy policy': 50.0, 'Platform interaction': 0.0,..}
            print('--->',category_wise_rating)

            risk_rating = self.calculate_risk_rating(option_wise_count['Pass'], option_wise_count['Fail'])
            
            # save risk_rating and category_wise_rating to DB
            data = dict(category_wise_rating=category_wise_rating, risk_rating=risk_rating, option_wise_count=option_wise_count)
            print(app_detail[0],'data---->',data)
            serializer = AppDetailSerializer(app_detail[0], data=data, partial=True)
            if serializer.is_valid():
                serializer.save()

                # change app status to 'submitted_to_rev'
                app_info = AppInfo.objects.filter(id=app_detail[0].app_info_id)
                app_status = 'submitted_to_rev'
                status_explanation = 'Application submitted to Reviewer.'
                todayDate = date.today().isoformat()
                action = "[{}] {} ({})".format(current_user.role, current_user.full_name, user_email)
                timeline_data = {'status_explanation': status_explanation, 'date': todayDate, 'action_by': action}
                timeline = app_info[0].timeline
                timeline.append(timeline_data)
                data = dict(status=app_status, timeline=timeline) 
                serializer = AppInfoSerializer(app_info[0],data=data,partial=True)
                if serializer.is_valid():
                    serializer.save()
                
            #send response
            responseData = { 'category_wise_rating': category_wise_rating,
                                'option_wise_count': option_wise_count,
                                'risk_rating': risk_rating}
            return Response({'status_code':200, 'error':[],'payload': responseData},status=status.HTTP_200_OK)   
        else:
            return Response({'status_code':404, 'message':'Invalid App Detail Id'}, status=status.HTTP_400_BAD_REQUEST)

        #response: 
        # {
            # option_wise_count: {
            #     pass: 1,
            #     fail: 3,
            #     NA: 4,
            #     UTV: 2
            # },
            # category_wise_rating: {
            #     ..
            # },
            # risk_rating: 3
        # }

